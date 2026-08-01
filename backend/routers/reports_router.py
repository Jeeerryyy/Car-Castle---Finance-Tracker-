"""Activity log + reports (PDF + Excel) routes."""
import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from deps import get_db, get_current_user, require_super_admin

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import mm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/activity", tags=["activity"])


@router.get("")
async def list_activity(user: dict = Depends(require_super_admin),
                        limit: int = 200,
                        target_collection: Optional[str] = None,
                        admin_id: Optional[str] = None):
    db = get_db()
    q = {}
    if target_collection:
        q["target_collection"] = target_collection
    if admin_id:
        q["admin_id"] = admin_id
    docs = await db.activity_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return docs


# ---------- Reports ----------
reports_router = APIRouter(prefix="/reports", tags=["reports"])


def _fmt_inr(n: float) -> str:
    n = float(n or 0)
    # Indian numbering
    s = f"{n:,.0f}"
    return f"Rs {s}"


async def _gather_report(db, month: str):
    """month = YYYY-MM"""
    q = {"start_date": {"$regex": f"^{month}"}}
    bookings = await db.bookings.find(q, {"_id": 0}).sort("start_date", 1).to_list(2000)
    settings = await db.settings.find_one({"id": "default"}) or {}
    savings_pct = float(settings.get("savings_percent", 10))

    total_income = sum(float(b["customer_rate"]) for b in bookings)
    total_owner_cost = sum(float(b["cost_rate"]) for b in bookings)
    total_agent_fee = sum(float(b.get("agent_fee", 0)) for b in bookings)
    total_margin = sum(float(b["margin"]) for b in bookings)
    total_net = sum(float(b["net_profit"]) for b in bookings)
    savings = total_net * (savings_pct / 100.0)

    owners = await db.car_owners.find({}, {"_id": 0}).to_list(500)
    agents = await db.agents.find({}, {"_id": 0}).to_list(500)
    car_map = {c["id"]: c for c in await db.cars.find({}, {"_id": 0}).to_list(500)}
    for b in bookings:
        car = car_map.get(b["car_id"], {})
        b["car_model"] = car.get("model", "—")
        b["car_registration"] = car.get("registration_no", "—")

    return {
        "month": month,
        "bookings": bookings,
        "owners": owners,
        "agents": agents,
        "totals": {
            "income": total_income,
            "owner_cost": total_owner_cost,
            "agent_fee": total_agent_fee,
            "margin": total_margin,
            "net_profit": total_net,
            "savings": savings,
            "savings_percent": savings_pct,
        },
    }


@reports_router.get("/monthly.pdf")
async def monthly_pdf(month: str = Query(..., description="YYYY-MM"),
                      user: dict = Depends(require_super_admin)):
    db = get_db()
    data = await _gather_report(db, month)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=18 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    orange = colors.HexColor("#EA580C")
    slate900 = colors.HexColor("#0F172A")
    slate500 = colors.HexColor("#64748B")
    slate100 = colors.HexColor("#F1F5F9")

    title_style = ParagraphStyle("t", parent=styles["Title"],
                                 textColor=slate900, fontSize=22, spaceAfter=6)
    sub_style = ParagraphStyle("s", parent=styles["Normal"],
                               textColor=slate500, fontSize=10, spaceAfter=14)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"],
                        textColor=slate900, fontSize=13, spaceAfter=8, spaceBefore=16)
    brand = ParagraphStyle("brand", parent=styles["Normal"],
                           textColor=orange, fontSize=11, spaceAfter=2)

    story = []
    story.append(Paragraph("CAR CASTLE GOA", brand))
    story.append(Paragraph(f"Monthly Report — {data['month']}", title_style))
    story.append(Paragraph("Self-drive rentals · Owner & agent ledgers · Airport transfers", sub_style))

    t = data["totals"]
    kpi_data = [
        ["Total Bookings", "Total Income", "Owner Payables", "Agent Fees", "Net Margin", "Savings"],
        [str(len(data["bookings"])), _fmt_inr(t["income"]), _fmt_inr(t["owner_cost"]),
         _fmt_inr(t["agent_fee"]), _fmt_inr(t["net_profit"]),
         f"{_fmt_inr(t['savings'])} ({t['savings_percent']:.0f}%)"],
    ]
    kpi_table = Table(kpi_data, colWidths=[30 * mm] * 6)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate900),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, 1), (-1, 1), slate100),
        ("TEXTCOLOR", (4, 1), (5, 1), colors.HexColor("#047857")),
    ]))
    story.append(kpi_table)

    # Bookings table
    story.append(Paragraph("Bookings", h2))
    rows = [["Date", "Customer", "Car", "Owner Cost", "Customer", "Margin", "Net"]]
    for b in data["bookings"]:
        rows.append([
            b.get("start_date", "")[:10],
            b.get("customer_name", "")[:20],
            f"{b.get('car_model','')}",
            _fmt_inr(b["cost_rate"]),
            _fmt_inr(b["customer_rate"]),
            _fmt_inr(b["margin"]),
            _fmt_inr(b["net_profit"]),
        ])
    if len(rows) == 1:
        rows.append(["—", "No bookings this month", "", "", "", "", ""])
    bt = Table(rows, colWidths=[22 * mm, 32 * mm, 32 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm])
    bt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate100),
        ("TEXTCOLOR", (0, 0), (-1, 0), slate900),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
    ]))
    story.append(bt)

    # Payables
    story.append(Paragraph("Owner Payables (Outstanding)", h2))
    o_rows = [["Owner", "Contact", "Total Owed", "Paid", "Balance"]]
    for o in data["owners"]:
        bal = float(o["total_owed"]) - float(o["total_paid"])
        if bal > 0.01:
            o_rows.append([o["name"], o.get("contact", ""), _fmt_inr(o["total_owed"]),
                           _fmt_inr(o["total_paid"]), _fmt_inr(bal)])
    if len(o_rows) == 1:
        o_rows.append(["—", "All settled", "", "", ""])
    ot = Table(o_rows, colWidths=[45 * mm, 35 * mm, 30 * mm, 30 * mm, 30 * mm])
    ot.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate100),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("TEXTCOLOR", (4, 1), (4, -1), colors.HexColor("#B91C1C")),
    ]))
    story.append(ot)

    story.append(Paragraph("Agent Payables (Outstanding)", h2))
    a_rows = [["Agent", "Contact", "Total Owed", "Paid", "Balance"]]
    for a in data["agents"]:
        bal = float(a["total_owed"]) - float(a["total_paid"])
        if bal > 0.01:
            a_rows.append([a["name"], a.get("contact", ""), _fmt_inr(a["total_owed"]),
                           _fmt_inr(a["total_paid"]), _fmt_inr(bal)])
    if len(a_rows) == 1:
        a_rows.append(["—", "All settled", "", "", ""])
    at = Table(a_rows, colWidths=[45 * mm, 35 * mm, 30 * mm, 30 * mm, 30 * mm])
    at.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), slate100),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("TEXTCOLOR", (4, 1), (4, -1), colors.HexColor("#B91C1C")),
    ]))
    story.append(at)

    story.append(Spacer(1, 12))
    footer = ParagraphStyle("f", parent=styles["Normal"], textColor=slate500,
                            fontSize=8, alignment=1)
    story.append(Paragraph(f"Car Castle Goa · Confidential internal document · Generated for {month}",
                           footer))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="car-castle-goa-{month}.pdf"'},
    )


@reports_router.get("/monthly.xlsx")
async def monthly_xlsx(month: str = Query(..., description="YYYY-MM"),
                       user: dict = Depends(require_super_admin)):
    db = get_db()
    data = await _gather_report(db, month)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    right = Alignment(horizontal="right")

    def _style_header(ws, row):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Car Castle Goa — Monthly Report"])
    ws["A1"].font = Font(bold=True, size=16, color="EA580C")
    ws.append([f"Month: {data['month']}"])
    ws.append([])
    ws.append(["Metric", "Value (INR)"])
    _style_header(ws, ws.max_row)
    t = data["totals"]
    for label, val in [
        ("Bookings", len(data["bookings"])),
        ("Total Income", t["income"]),
        ("Owner Payables (cost)", t["owner_cost"]),
        ("Agent Fees", t["agent_fee"]),
        ("Total Margin", t["margin"]),
        ("Net Profit (after agent fees)", t["net_profit"]),
        (f"Savings ({t['savings_percent']:.0f}%)", t["savings"]),
    ]:
        ws.append([label, val])
    for col in ("A", "B"):
        ws.column_dimensions[col].width = 32

    # Bookings sheet
    ws2 = wb.create_sheet("Bookings")
    ws2.append(["Date", "Customer", "Car", "Reg", "Owner Cost", "Customer Rate",
                "Margin", "Agent Fee", "Net Profit", "Status", "Transfer"])
    _style_header(ws2, 1)
    for b in data["bookings"]:
        ws2.append([
            b.get("start_date", "")[:10], b.get("customer_name", ""),
            b.get("car_model", ""), b.get("car_registration", ""),
            float(b["cost_rate"]), float(b["customer_rate"]),
            float(b["margin"]), float(b.get("agent_fee", 0)),
            float(b["net_profit"]), b.get("status", ""),
            b.get("transfer_type", "none"),
        ])
    for c in range(5, 10):
        for row in ws2.iter_rows(min_row=2, min_col=c, max_col=c):
            for cell in row:
                cell.alignment = right
    for col_letter in "ABCDEFGHIJK":
        ws2.column_dimensions[col_letter].width = 16

    # Owner Payables
    ws3 = wb.create_sheet("Owner Payables")
    ws3.append(["Owner", "Contact", "Total Owed", "Total Paid", "Balance"])
    _style_header(ws3, 1)
    for o in data["owners"]:
        ws3.append([o["name"], o.get("contact", ""), float(o["total_owed"]),
                    float(o["total_paid"]), float(o["total_owed"]) - float(o["total_paid"])])
    for col_letter in "ABCDE":
        ws3.column_dimensions[col_letter].width = 22

    # Agent Payables
    ws4 = wb.create_sheet("Agent Payables")
    ws4.append(["Agent", "Contact", "Total Owed", "Total Paid", "Balance"])
    _style_header(ws4, 1)
    for a in data["agents"]:
        ws4.append([a["name"], a.get("contact", ""), float(a["total_owed"]),
                    float(a["total_paid"]), float(a["total_owed"]) - float(a["total_paid"])])
    for col_letter in "ABCDE":
        ws4.column_dimensions[col_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="car-castle-goa-{month}.xlsx"'},
    )
